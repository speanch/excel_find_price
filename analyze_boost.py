#!/usr/bin/env python3
"""
Декодирование и сравнение выгрузок «Эластичный бустинг» Ozon (products-*.xlsx).

Столбцы BC–BI в листе «Товары и цены» — это XOR-замаскированные копии значений,
которые уже есть в открытых столбцах. Декодирование:
    реальное_значение = BITXOR(BC, X) / 1000

Расшифровка:
    BC — ключ (случайная соль 0..999, новая при каждой выгрузке)
    BD → минимальный бустинг + 1   (1.15 → +15%)
    BE → максимальный бустинг + 1  (1.75 → +75%)
    BF → Q, «Цена для максимального бустинга»
    BG → P, «Цена для минимального бустинга»
    BH → базовая цена, при которой рассчитан текущий бустинг (= M или I)
    BI → текущий бустинг + 1

Сам бустинг (столбец O) считается по формуле и равен 0..0.75 (то есть +0%..+75%).

Использование:
    python analyze_boost.py *.xlsx
    python analyze_boost.py *.xlsx --out boost_compare.csv
"""
import argparse
import csv
import glob
import os
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SHEET = "Товары и цены"
HEADER_ROW = 2
DATA_START = 4


def xor(a, b):
    if a is None or b is None:
        return None
    return int(a) ^ int(b)


def decode(bc, value):
    """BITXOR(BC, value) / 1000."""
    if bc is None or value is None:
        return None
    return xor(bc, value) / 1000.0


def to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def compute_boost_pct(M, P, Q, L, bd, be, bf, bg, bh, bi):
    """Восстанавливает бустинг (%) по формуле столбца O (проверено 1174/1174)."""
    if L in (None, ""):
        return None
    M = to_float(M)
    P = to_float(P)
    if M is None or P is None:
        return None
    if M == bh:
        return round((bi - 1.0) * 100, 3)
    if M > P:
        return 0.0
    if (be < bd) or (bf > bg):
        return round((be - 1.0) * 100, 3)
    if bf == 0:
        return round((bd - 1.0) * 100, 3)
    if M <= bf:
        return round((be - 1.0) * 100, 3)
    if M >= bg:
        return round((bd - 1.0) * 100, 3)
    interp = bd + ((bg - M) / (bg - bf)) * (be - bd)
    return round((interp - 1.0) * 100, 3)


def load_file(path):
    wb = load_workbook(path, data_only=True)
    if SHEET not in wb.sheetnames:
        raise ValueError(f"{path}: нет листа «{SHEET}»")
    ws = wb[SHEET]
    cols = {get_column_letter(c): c for c in range(1, ws.max_column + 1)}

    def cell(letter, r):
        c = cols.get(letter)
        return ws.cell(row=r, column=c).value if c else None

    records = {}
    for r in range(DATA_START, ws.max_row + 1):
        ozon_id = cell("A", r)
        if ozon_id is None:
            continue
        bc = cell("BC", r)
        if bc is None:
            continue
        bd = decode(bc, cell("BD", r))
        be = decode(bc, cell("BE", r))
        bf = decode(bc, cell("BF", r))
        bg = decode(bc, cell("BG", r))
        bh = decode(bc, cell("BH", r))
        bi = decode(bc, cell("BI", r))
        M = cell("M", r)
        L = cell("L", r)
        art = str(cell("C", r) or "").strip()
        sku = str(cell("B", r) or "").strip()
        name = str(cell("E", r) or "").strip()
        P = cell("P", r)
        Q = cell("Q", r)
        I = cell("I", r)
        boost = compute_boost_pct(M, P, Q, L, bd, be, bf, bg, bh, bi)
        records[art] = {
            "art": art,
            "ozon_id": ozon_id,
            "sku": sku,
            "name": name,
            "L": L,
            "M": to_float(M),
            "I": to_float(I),
            "P": to_float(P),
            "Q": to_float(Q),
            "boost_pct": boost,
            "min_boost_pct": round((bd - 1.0) * 100, 3) if bd is not None else None,
            "max_boost_pct": round((be - 1.0) * 100, 3) if be is not None else None,
            "bh": bh,
        }
    return records


def short_label(path):
    base = os.path.basename(path)
    stem = os.path.splitext(base)[0]
    for token in stem.replace("(", " ").replace(")", " ").split():
        if token.isdigit():
            return token
    return stem


def main():
    ap = argparse.ArgumentParser(description="Сравнение выгрузок эластичного бустинга Ozon")
    ap.add_argument("files", nargs="+", help="xlsx-файлы выгрузок")
    ap.add_argument("--out", default=None, help="путь к CSV-отчёту")
    args = ap.parse_args()

    files = []
    for pat in args.files:
        files.extend(glob.glob(pat))
    files = sorted(set(files))
    if len(files) < 1:
        print("Не найдено файлов", file=sys.stderr)
        sys.exit(1)

    labels = [short_label(f) for f in files]
    data = {}
    for f, lab in zip(files, labels):
        data[lab] = load_file(f)
        print(f"{f}  →  {len(data[lab])} товаров")

    arts = sorted(set().union(*(set(d) for d in data.values())))

    # ── изменения по каждому артикулу ────────────────────────────────────
    def series(field):
        return [data[lab].get(art, {}).get(field) for lab in labels]

    changed = []
    for art in arts:
        boosts = series("boost_pct")
        ps = series("P")
        qs = series("Q")
        ms = series("M")
        ls = series("L")
        present = [i for i, b in enumerate(boosts) if b is not None]
        if len(present) < 2:
            continue
        first, last = present[0], present[-1]
        d_boost = (boosts[last] or 0) - (boosts[first] or 0)
        if (
            d_boost != 0
            or ps[first] != ps[last]
            or qs[first] != qs[last]
            or ms[first] != ms[last]
        ):
            rec = data[labels[last]][art]
            changed.append(
                {
                    "art": art,
                    "name": rec.get("name", ""),
                    "boost_first": boosts[first],
                    "boost_last": boosts[last],
                    "d_boost": d_boost,
                    "p_first": ps[first],
                    "p_last": ps[last],
                    "q_first": qs[first],
                    "q_last": qs[last],
                    "m_first": ms[first],
                    "m_last": ms[last],
                }
            )

    # сортировка: сначала у кого изменился бустинг, потом по модулю дельты
    changed.sort(key=lambda x: (x["d_boost"] == 0, -abs(x["d_boost"])))

    print(f"\nИзменения между «{labels[0]}» и «{labels[-1]}»: {len(changed)} товаров")
    print(
        f"{'Артикул':<16} {'Название':<40} "
        f"{'бустинг':>8} {'Δ':>8} | {'P мин':>9} {'P→':>9} | {'Q макс':>9} {'Q→':>9} | {'M цена':>9} {'M→':>9}"
    )
    for c in changed[:60]:
        name = c["name"][:38] if c["name"] else ""
        print(
            f"{c['art']:<16} {name:<40} "
            f"{c['boost_last']:>7.1f}% {c['d_boost']:>+7.1f} | "
            f"{c['p_first']:>9} {c['p_last']:>9} | "
            f"{c['q_first']:>9} {c['q_last']:>9} | "
            f"{c['m_first']:>9} {c['m_last']:>9}"
        )
    if len(changed) > 60:
        print(f"… и ещё {len(changed) - 60} товаров (см. CSV)")

    # ── сводка ────────────────────────────────────────────────────────────
    print("\nСводка:")
    up = sum(1 for c in changed if c["d_boost"] > 0)
    down = sum(1 for c in changed if c["d_boost"] < 0)
    same = sum(1 for c in changed if c["d_boost"] == 0)
    print(f"  бустинг вырос: {up}, упал: {down}, не изменился (но изменились цены): {same}")

    # ── CSV ──────────────────────────────────────────────────────────────
    if args.out:
        with open(args.out, "w", newline="", encoding="utf-8-sig") as fh:
            w = csv.writer(fh, delimiter=";")
            header = ["Артикул", "OzonID", "SKU", "Название"]
            for lab in labels:
                header += [
                    f"бустинг%_{lab}",
                    f"P_мин_{lab}",
                    f"Q_макс_{lab}",
                    f"M_цена_{lab}",
                    f"I_тек_{lab}",
                    f"участие_{lab}",
                ]
            header += ["Δ_бустинг"]
            w.writerow(header)
            for art in arts:
                rec = next(
                    (data[lab].get(art) for lab in labels if art in data[lab]), {}
                )
                row = [
                    art,
                    rec.get("ozon_id", ""),
                    rec.get("sku", ""),
                    rec.get("name", ""),
                ]
                for lab in labels:
                    d = data[lab].get(art, {})
                    row += [
                        d.get("boost_pct"),
                        d.get("P"),
                        d.get("Q"),
                        d.get("M"),
                        d.get("I"),
                        d.get("L"),
                    ]
                boosts = series("boost_pct")
                present = [i for i, b in enumerate(boosts) if b is not None]
                d_boost = (
                    (boosts[present[-1]] or 0) - (boosts[present[0]] or 0)
                    if len(present) >= 2
                    else ""
                )
                row.append(d_boost)
                w.writerow(row)
        print(f"\nCSV сохранён: {args.out}")


if __name__ == "__main__":
    main()
